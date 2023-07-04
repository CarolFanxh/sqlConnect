from Config import login
import openpyxl
import datetime
from Config import filePath
import json
import os

#
#
# def escape_quotes(json_str):
#     # 将双引号替换为反斜杠和双引号
#     escaped_str = json_str.replace('"', '\\"')
#     return escaped_str

# 测试代码
# json_str = '{"keyid":"5e1986a6-00bc-451c-ab4f-cfebd3d7b59b","fkeyid":"05390d34-bac3-4663-a423-a3453c568215"}'
# escaped_str = escape_quotes(json_str)
# print(escaped_str)

bl_no = 'ZGSHA0382001043'
sqlTDev = f"select json_data from callback_in where json_data like '%{bl_no}%' and source = 'YUNDANG'"
print(sqlTDev)
sqlTDevResult = login.pull_jira_data(sqlTDev, '生产-pub-pub-mdm-读写', 'mdm')
print(sqlTDevResult)

now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
提取数据库json_data的数据存入excel
 # 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
 # 获取第一个工作表
ws = wb.active
 # 定义要提取的数据
# sqlTDevResult = [[{'a'}],[{'b'}],[{'c'}]]
 # 遍历数据并将其写入Excel
for i, row in enumerate(sqlTDevResult):
    for j, cell in enumerate(row):
        # 获取大括号里面的内容
        if isinstance(cell, str):
            content = cell
        else:
            content = cell.pop()
        # 将内容写入Excel
        print(content)
        ws.cell(row=i+1, column=j+1, value=content)

# 删除以前的同名Excel文件
# for file in os.listdir():
#     if file.startswith("sqlResult_"):
#         os.remove(file)

 # 保存Excel文件
curDir = filePath.get_current_directory()
excelPath = os.path.join(curDir, 'sqlResultData')
print(excelPath)
wb.save(os.path.join(excelPath, f"sqlResultData{now}.xlsx"))

#提取数据库json_data的数据，转义后存入excel
 # 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
 # 获取第一个工作表
ws = wb.active
 # 遍历数据并将其写入Excel
for i, row in enumerate(sqlTDevResult):
    for j, cell in enumerate(row):
        # 获取大括号里面的内容
        if isinstance(cell, str):
            data = json.dumps(cell)#转义引号
        else:
            data = cell.pop()
        # 将内容写入Excel
        # print(data)
        ws.cell(row=i+1, column=j+1, value=data)

# 删除以前的同名Excel文件
# for file in os.listdir():
#     if file.startswith("post_sqlResult"):
#         os.remove(file)

# 保存Excel文档
curDir = filePath.get_current_directory()
excelPath2 = os.path.join(curDir, 'postSqlResultData')
print(excelPath2)
wb.save(os.path.join(excelPath2, f"postSqlResult_{now}.xlsx"))
# dir_path = os.path.join('', '../sqlResultData')
# if not os.path.exists(dir_path):
#     os.makedirs(dir_path)
# file_path = os.path.join(dir_path, f"post_sqlResult_{now}.xlsx")
# wb.save(file_path)
